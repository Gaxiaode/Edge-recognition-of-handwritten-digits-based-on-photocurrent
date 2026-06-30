import os
import cv2
import torch
import numpy as np
from openpyxl import load_workbook

from PIL import Image


class BrightnessTransform:
    def __init__(self, brightness=1.0, return_type="pil"):
        """
        brightness: 亮度衰减系数，例如 1.0, 0.6, 0.2
        return_type:
            'pil'   -> 返回 PIL Image, 方便后面接 transforms.ToTensor()
            'numpy' -> 返回 NumPy array
            'tensor'-> 返回 Tensor, [1, H, W]，范围 [0, 1]
        """
        self.brightness = brightness
        self.return_type = return_type

    def __call__(self, img):
        # 1. 根据输入类型，统一转成 NumPy
        if isinstance(img, Image.Image):
            # PIL Image -> NumPy
            img_np = np.array(img.convert("L"))

        elif isinstance(img, np.ndarray):
            # NumPy 输入
            img_np = img

            # # # 如果是 RGB 图像，转灰度
            # # if img_np.ndim == 3:
            # #     # 简单 RGB -> Gray
            # #     img_np = 0.299 * img_np[:, :, 0] + 0.587 * img_np[:, :, 1] + 0.114 * img_np[:, :, 2]

        elif torch.is_tensor(img):
            # Tensor 输入
            img_tensor = img.detach().cpu()

            # [1, H, W] -> [H, W]
            if img_tensor.ndim == 3 and img_tensor.shape[0] == 1:
                img_np = img_tensor.squeeze(0).numpy()

            # [H, W]
            elif img_tensor.ndim == 2:
                img_np = img_tensor.numpy()

            else:
                raise ValueError(f"Unsupported tensor shape: {img_tensor.shape}")

        else:
            raise TypeError(f"Unsupported image type: {type(img)}")

        # 2. 统一转成 float32
        img_np = img_np.astype(np.float32)

        # 3. 如果输入是 [0, 1]，先变成 [0, 255]
        if img_np.max() <= 1.0:
            img_np = img_np * 255.0

        # 4. 亮度衰减
        img_np = img_np * self.brightness

        # 5. 限制范围到 [0, 255]
        img_np = np.clip(img_np, 0, 255).astype(np.uint8)

        # 6. 根据需要返回不同格式
        if self.return_type == "pil":
            return Image.fromarray(img_np)

        elif self.return_type == "numpy":
            return img_np

        elif self.return_type == "tensor":
            img_np = img_np.astype(np.float32) / 255.0
            return torch.from_numpy(img_np).unsqueeze(0)

        else:
            raise ValueError("return_type must be 'pil', 'numpy', or 'tensor'")


class EdgeTransform:
    def __init__(self, use=True, method="canny"):
        """
        method: 'sobel', 'prewitt', 'canny'
        brightness: 亮度衰减系数，例如 1.0, 0.6, 0.2
        """
        self.use = use
        self.method = method


    def __call__(self, img):
        """
        img: torchvision 读出来的 PIL Image
        return: Tensor, shape = [1, 28, 28]
        """
        # 1. PIL Image -> NumPy 灰度图
        if type(img) is not np.ndarray:
            img = np.array(img.convert("L"))  # shape: [28, 28], uint8, 0~255

        # 2. 限制范围
        img = img.astype(np.float32)
        img = np.clip(img, 0, 255).astype(np.uint8)

        if not self.use:
            return img

        # 3. 边缘检测
        if self.method == "sobel":
            edge = self.sobel_edge(img)

        elif self.method == "prewitt":
            edge = self.prewitt_edge(img)

        elif self.method == "canny":
            edge = cv2.Canny(img, threshold1=50, threshold2=150)

        else:
            raise ValueError("method must be 'sobel', 'prewitt', or 'canny'")

        # 4. 归一化到 [0, 1]
        edge = edge.astype(np.float32) / 255.0      # 因为这里的输出默认是np.ndarray ToTensor不会进行归一化

        # # 5. NumPy -> Tensor，并增加通道维度
        # edge = torch.from_numpy(edge).unsqueeze(0)  # [28, 28] -> [1, 28, 28]

        return edge

    def sobel_edge(self, img):
        gx = cv2.Sobel(img, cv2.CV_32F, 1, 0, ksize=3)
        gy = cv2.Sobel(img, cv2.CV_32F, 0, 1, ksize=3)

        edge = np.sqrt(gx ** 2 + gy ** 2)

        # 归一化到 0~255
        edge = cv2.normalize(edge, None, 0, 255, cv2.NORM_MINMAX)
        edge = edge.astype(np.uint8)

        return edge

    def prewitt_edge(self, img):
        kernel_x = np.array([
            [-1, 0, 1],
            [-1, 0, 1],
            [-1, 0, 1]
        ], dtype=np.float32)

        kernel_y = np.array([
            [-1, -1, -1],
            [0, 0, 0],
            [1, 1, 1]
        ], dtype=np.float32)

        gx = cv2.filter2D(img, cv2.CV_32F, kernel_x)
        gy = cv2.filter2D(img, cv2.CV_32F, kernel_y)

        edge = np.sqrt(gx ** 2 + gy ** 2)

        edge = cv2.normalize(edge, None, 0, 255, cv2.NORM_MINMAX)
        edge = edge.astype(np.uint8)

        return edge
    

class RaplaceTransform:
    def __init__(self, args):
        self.light_path = args.light_path  # 光电流转换路径
        self.num_groups = 10
        self.return_original_range = getattr(args, "return_original_range", True)
        self.current_groups = self._load_current_groups()
        self.current_min, self.current_max = self._get_current_range()
        self.use_edge = args.use_edge
        # light_path是 ./dataset/light
        # 里面有用的文件格式是 0.0x_large_current.xlsx
        '''
        需要将每一个文件里面的光电流读出来，电流信息在表格第二列（第一列是时间）
        一共有10组数据，都要保存好
        然后在 __call__中，首先需要检查，这个图像的范围是否在0到1，比方说输入的是
        0到255范围，那么需要先归一化，再去做替换，然后再返回0到255

        这个替换方法如下：归一化后图像中的某个像素点，假设像素值是0.35，那么就要从
        0.04_large_current.xlsx所得值中随机挑一个替换，注意！是随机挑一个！
        如果是0.95，就是从0.1_large_current.xlsx中随机挑一个做替换
        格式是nd.array
        '''

    def __call__(self, img):
        img_np = self._to_numpy(img).astype(np.float32)
        if img_np.size == 0:
            raise ValueError("Input image is empty.")
        if not np.isfinite(img_np).all():
            raise ValueError("Input image contains NaN or infinite values.")
        if img_np.min() < 0:
            raise ValueError("Input image values must be non-negative.")

        scale = 1.0
        if img_np.max() > 1.0:
            scale = 255.0 if img_np.max() <= 255.0 else float(img_np.max())

        img_norm = np.clip(img_np / scale, 0.0, 1.0)
        group_indices = np.ceil(img_norm * self.num_groups).astype(np.int64)
        group_indices = np.clip(group_indices, 1, self.num_groups)

        replaced = np.empty_like(img_norm, dtype=np.float32)
        for group_idx, current_values in self.current_groups.items():
            mask = group_indices == group_idx
            count = int(mask.sum())
            if count == 0:
                continue

            random_indices = np.random.randint(0, len(current_values), size=count)
            replaced[mask] = current_values[random_indices]

        replaced = np.abs(replaced)
        replaced = self._map_current_to_255(replaced)
        replaced = np.clip(replaced, 0, 255).astype(np.uint8)

        return replaced
        # return Image.fromarray(replaced, mode="L")

    def _load_current_groups(self):
        current_groups = {}
        for group_idx in range(1, self.num_groups + 1):
            file_name = f"{group_idx / 100:g}_large_current.xlsx"
            file_path = os.path.join(self.light_path, file_name)
            current_values = self._read_current_file(file_path)

            if current_values.size == 0:
                raise ValueError(f"No valid current values found in {file_path}.")

            current_groups[group_idx] = np.abs(current_values)

        return current_groups

    def _get_current_range(self):
        all_currents = np.concatenate(list(self.current_groups.values()))
        current_min = float(all_currents.min())
        current_max = float(all_currents.max())
        if current_max <= current_min:
            raise ValueError("Current values must have a non-zero range.")
        return current_min, current_max

    def _map_current_to_255(self, current_values):
        mapped = (current_values - self.current_min) / (self.current_max - self.current_min)
        mapped = np.clip(mapped, 0.0, 1.0) * 255.0
        return mapped

    def _read_current_file(self, file_path):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Current file not found: {file_path}")

        values = []
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                if len(row) < 2 or row[1] is None:
                    continue

                try:
                    values.append(float(row[1]))
                except (TypeError, ValueError):
                    continue
        finally:
            workbook.close()

        return np.asarray(values, dtype=np.float32)

    def _to_numpy(self, img):
        if isinstance(img, Image.Image):
            return np.asarray(img.convert("L"))

        if isinstance(img, np.ndarray):
            img_np = img
        elif torch is not None and torch.is_tensor(img):
            img_np = img.detach().cpu().numpy()
        else:
            raise TypeError(f"Unsupported image type: {type(img)}")

        if img_np.ndim == 2:
            return img_np
        if img_np.ndim == 3 and img_np.shape[0] == 1:
            return img_np.squeeze(0)
        if img_np.ndim == 3 and img_np.shape[-1] == 1:
            return img_np.squeeze(-1)
        if img_np.ndim == 3 and img_np.shape[-1] in (3, 4):
            rgb = img_np[..., :3].astype(np.float32)
            return 0.299 * rgb[..., 0] + 0.587 * rgb[..., 1] + 0.114 * rgb[..., 2]

        raise ValueError(f"Unsupported image shape: {img_np.shape}")

